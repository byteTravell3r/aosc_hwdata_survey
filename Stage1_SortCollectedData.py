import re
import time
import openpyxl
from hwdata_checker import update_hwdata, PCI, USB
from utils import *

# 更新 pci.ids 和 usb.ids
print_info("========================================")
print_info("正在获取 hwdata 设备数据库...")
if not update_hwdata():
    exit(1)

# hwdata 问卷信息汇总表(需要手动下载!)
hwdata_excel_file = "hwdata.xlsx"

# 用来判别每行是否为有效信息的关键词列表
keyword_list = [
    "Subsystem:",
    "Kernel modules:",
    "Kernel driver in use:",
    "Flags:",
    "Capabilities:",
    "DeviceName:",
    "None",
    "pcilib:",
    "lspci:",
    "$"
]


# 判断一行文本是否可被直接排除
def line_is_useful(input_text: str) -> bool:
    if input_text == "":
        return False
    if input_text[0] == "#":
        return False
    # 剔除无关信息和可能的杂项
    for word in keyword_list:
        if input_text.find(word) != -1:
            return False
    return True


# 通过文本行开头的"设备位置"部分判断其是否为有效的PCIe设备信息行
def pcie_input_is_valid(input_text: str) -> bool:
    pattern_type1 = re.compile(r"[A-Za-z0-9]+:[A-Za-z0-9]+\.[A-Za-z0-9]+", re.IGNORECASE)
    match_type1 = pattern_type1.match(input_text)
    pattern_type2 = re.compile(r"[A-Za-z0-9]+:[A-Za-z0-9]+:[A-Za-z0-9]+\.[A-Za-z0-9]+", re.IGNORECASE)
    match_type2 = pattern_type2.match(input_text)
    if match_type1 or match_type2:
        return True
    else:
        return False


# 从有效行中提取 PCIe 设备编号
def strip_pcie_id(input_text: str) -> str:
    pattern = r'[0-9a-f]{4}:[0-9a-f]{4}'
    result = re.findall(pattern, input_text, flags=re.IGNORECASE)
    if result:
        return result[0]
    else:
        # 从行中拆出有用的部分
        index = input_text.index(":", 8)
        input_text = input_text[index + 2:]

        # 如果原始数据行中只有 PID 而没有 VID, 则尝试将其补全 (有部分数据的确如此)
        pattern = r'[0-9a-f]{4}'
        result = re.findall(pattern, input_text, flags=re.IGNORECASE)
        if result:
            if input_text.find("[AMD]") != -1:
                return f"1022:{result[0]}"
            if input_text.find("[AMD/ATI]") != -1:
                return f"1002:{result[0]}"
            if input_text.find("Intel") != -1:
                return f"8086:{result[0]}"
            if input_text.find("Loongson") != -1:
                return f"0014:{result[0]}"
            return ""
        else:
            print_warn("未找到该行的有效设备编号: " + input_text)
            return ""

# 从有效行中提取 USB 设备编号
def strip_usb_id(input_text: str) -> str:
    pattern = r'[0-9a-f]{4}:[0-9a-f]{4}'
    result = re.findall(pattern, input_text, flags=re.IGNORECASE)
    if result:
        return result[0]
    else:
        return ""

# 新建列表和集合
pcie_useful_data_set = set()
pcie_not_useful_data_set = set()
pcie_already_existed_data_set = set()
pcie_not_recognizable_data_list = list()

usb_useful_data_set = set()
usb_already_existed_data_set = set()
usb_not_useful_data_set = set()

# 打开文件并读取数据
hwdata_workbook = openpyxl.load_workbook(hwdata_excel_file)
hwdata_raw_sheet = hwdata_workbook.worksheets[0]
pcie_raw_data_list = list(hwdata_raw_sheet.columns)[3]
usb_raw_data_list = list(hwdata_raw_sheet.columns)[4]
hwdata_workbook.close()

# ==================== PCIe 部分 ====================

# 遍历列表, 逐行判断是否为有效的 PCI(e) 数据
print_info("========================================")
print_info("正在整理 PCI(e) 设备信息...")
for block in pcie_raw_data_list:
    current_block = str(block.value).splitlines()
    for line in current_block:
        line = line.strip()
        if not line_is_useful(line):
            pcie_not_useful_data_set.add(line)
        else:
            if pcie_input_is_valid(line):
                id_data = strip_pcie_id(line)
                if id_data is not None:
                    pcie_useful_data_set.add(id_data)
            else:
                pcie_not_recognizable_data_list.append(line)
print_info("========================================")

# 对"无效信息"进行检查
for line in pcie_not_useful_data_set:
    if pcie_input_is_valid(line):
        pcie_useful_data_set.add(strip_pcie_id(line))

# 移除无关元素
try:
    pcie_useful_data_set.remove("")
    pcie_not_recognizable_data_list.remove("PCIe_PCI设备信息缺失项")
except KeyError:
    pass

# 查询该设备信息是否已经存在于pci.ids中, 如果是, 则记录
pci = PCI()
for device in pcie_useful_data_set:
    ret = pci.get_device(device[0:4], device[5:9])
    if ret:
        pcie_already_existed_data_set.add(device)

# 对两个集合取差集, 获得目前信息未知的设备编号
pcie_useful_data_set = pcie_useful_data_set.difference(pcie_already_existed_data_set)

# ==================== USB 部分 ====================

# 遍历列表,逐行判断是否为有效的 USB 数据
print_info("正在整理 USB 设备信息...")
for block in usb_raw_data_list:
    current_block = str(block.value).splitlines()
    for line in current_block:
        line = line.strip()
        if not line_is_useful(line):
            usb_not_useful_data_set.add(line)
        else:
            id_data = strip_usb_id(line)
            usb_useful_data_set.add(id_data)
print_info("========================================")

# 移除无关元素
try:
    usb_useful_data_set.remove("")
except KeyError:
    pass

# 查询该设备信息是否已经存在于usb.ids中, 如果是, 则记录
usb = USB()
for device in usb_useful_data_set:
    ret = usb.get_device(device[0:4], device[5:9])
    if ret:
        usb_already_existed_data_set.add(device)

# 对两个集合取差集, 获得目前信息未知的设备编号
usb_useful_data_set = usb_useful_data_set.difference(usb_already_existed_data_set)

# ==================== 汇总部分 ====================

# 按字母排序整理信息列表
pcie_useful_data_list = sorted(list(pcie_useful_data_set))
usb_useful_data_list = sorted(list(usb_useful_data_set))

# 保存到工作簿中:
hwdata_workbook = openpyxl.load_workbook(hwdata_excel_file)

try:
    del hwdata_workbook["整理后的PCIe数据"]
    del hwdata_workbook["整理后的USB数据"]
except KeyError:
    pass
sheet_count = len(hwdata_workbook.sheetnames)
hwdata_sorted_pcie_sheet = hwdata_workbook.create_sheet("整理后的PCIe数据", sheet_count)
hwdata_sorted_usb_sheet = hwdata_workbook.create_sheet("整理后的USB数据", sheet_count + 1)
for i in range(len(pcie_useful_data_list)):
    hwdata_sorted_pcie_sheet.cell(i + 1, 1, pcie_useful_data_list[i])
for i in range(len(usb_useful_data_list)):
    hwdata_sorted_usb_sheet.cell(i + 1, 1, usb_useful_data_list[i])
try:
    hwdata_workbook.save(hwdata_excel_file)
except PermissionError:
    print_warn("无法写入表格文件, 程序无法继续.")
    exit(1)
hwdata_workbook.close()

# 汇报结果
pcie_count_useful = len(pcie_useful_data_set)
pcie_count_total = pcie_count_useful + len(pcie_already_existed_data_set)

usb_count_useful = len(usb_useful_data_set)
usb_count_total = usb_count_useful + len(usb_already_existed_data_set)
print_info(f"整理完毕, 已将信息保存到表格中.")
print_info(f"收集表中有 {pcie_count_total} 条 PCI(e) 设备 ID, 其中有 {pcie_count_useful} 条是目前未知的.")
print_info(f"收集表中有 {usb_count_total} 条 USB 设备 ID, 其中有 {usb_count_useful} 条是目前未知的.")
print_data("========================================")
print_data("注意: 还有一些内容无法识别, 请检查:")
time.sleep(3)
print_data("PCI(e) 部分:")
print_data("========================================")
for line in pcie_not_recognizable_data_list:
    print_data(line)
print_data("========================================")
print_data("USB 部分:")
print_data("========================================")
for line in usb_not_useful_data_set:
    print_data(line)
print_data("========================================")

time.sleep(3)
print_info(
    "接下来将打开表格, 您需要检视和修改 \"hwdata.xlsx\"\n" +
    "确保所有的行都符合 [xxxx:xxxx] 的格式.\n" +
    "在完成修改后, 您可继续执行本程序."
)

wait_for_q()
open_file(hwdata_excel_file)