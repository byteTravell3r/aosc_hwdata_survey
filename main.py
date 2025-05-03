import re
import time
from concurrent.futures import ThreadPoolExecutor

import openpyxl

from hwdata_checker import update_hwdata, PCI, USB
from online_info_checker import *

# 更新 pci.ids 和 usb.ids
clear_screen()
print_info("在开始之前, 请确保您已经导出 \"hwdata 信息缺漏征集表\" 并以 xls(x) 格式保存到程序目录下!")
print_info("========================================")
print_info("正在获取/更新本地的 hwdata 设备数据库...")
if not update_hwdata():
    exit(1)

# hwdata 问卷信息汇总表(需要手动下载!)
hwdata_excel_file = "hwdata.xlsx"

# 用来判别每行是否为有效信息的关键词列表
# (如果具有以下关键词,则视为该行信息无效)

keyword_list = [
    "Subsystem:",
    "Kernel modules:",
    "Kernel driver in use:",
    "Flags:",
    "Capabilities:",
    "DeviceName:",
    "None",
    "pcilib:",
    "lspci:",
    "https"
    "$"
]


# 判断一行文本是否可被直接排除
def line_is_useful(input_text: str) -> bool:
    if input_text == "":
        return False
    if input_text[0] == "#":
        return False
    # 剔除无关信息和可能的杂项
    for word in keyword_list:
        if input_text.find(word) != -1:
            return False
    return True


# 通过文本行开头的"设备位置"部分判断其是否为有效的PCIe设备信息行
def pcie_input_is_valid(input_text: str) -> bool:
    pattern_type1 = re.compile(r"[A-Za-z0-9]+:[A-Za-z0-9]+\.[A-Za-z0-9]+", re.IGNORECASE)
    match_type1 = pattern_type1.match(input_text)
    pattern_type2 = re.compile(r"[A-Za-z0-9]+:[A-Za-z0-9]+:[A-Za-z0-9]+\.[A-Za-z0-9]+", re.IGNORECASE)
    match_type2 = pattern_type2.match(input_text)
    if match_type1 or match_type2:
        return True
    else:
        return False


# 从有效行中提取 PCIe 设备编号, (格式 0000:0000)
def strip_pcie_id(input_text: str) -> str:
    pattern = r'[0-9a-f]{4}:[0-9a-f]{4}'
    result = re.findall(pattern, input_text, flags=re.IGNORECASE)
    if result:
        return result[0]
    else:
        # 从行中拆出有用的部分
        index = input_text.index(":", 8)
        input_text = input_text[index + 2:]

        # 如果原始数据行中只有 PID 而没有 VID, 则尝试将其补全 (有部分数据的确如此)
        pattern = r'[0-9a-f]{4}'
        result = re.findall(pattern, input_text, flags=re.IGNORECASE)
        if result:
            if input_text.find("[AMD/ATI]") != -1:
                return f"1002:{result[0]}"
            if input_text.find("[AMD]") != -1:
                return f"1022:{result[0]}"
            if input_text.find("Intel") != -1:
                return f"8086:{result[0]}"
            if input_text.find("Loongson") != -1:
                return f"0014:{result[0]}"
            return ""
        else:
            print_warn("未找到该行的有效设备编号: " + input_text)
            return ""


# 从有效行中提取 USB 设备编号, (格式 0000:0000)
def strip_usb_id(input_text: str) -> str:
    pattern = r'[0-9a-f]{4}:[0-9a-f]{4}'
    result = re.findall(pattern, input_text, flags=re.IGNORECASE)
    if result:
        return result[0]
    else:
        return ""


# 新建列表和集合
pcie_useful_data_set = set()
pcie_not_useful_data_set = set()
pcie_already_existed_data_set = set()
pcie_not_recognizable_data_list = list()

usb_useful_data_set = set()
usb_already_existed_data_set = set()
usb_not_useful_data_set = set()

# 打开文件并读取数据
hwdata_workbook = openpyxl.load_workbook(hwdata_excel_file)
hwdata_raw_sheet = hwdata_workbook.worksheets[0]
pcie_raw_data_list = list(hwdata_raw_sheet.columns)[3]
usb_raw_data_list = list(hwdata_raw_sheet.columns)[4]
hwdata_workbook.close()

# ==================== PCIe 部分 ====================

# 遍历列表, 逐行判断是否为有效的 PCI(e) 数据
print_info("========================================")
time.sleep(3)
clear_screen()
print_info("========================================")
print_info("正在整理 PCI(e) 设备信息...")
for block in pcie_raw_data_list:
    current_block = str(block.value).splitlines()
    for line in current_block:
        line = line.strip()
        if not line_is_useful(line):
            pcie_not_useful_data_set.add(line)
        else:
            if pcie_input_is_valid(line):
                id_data = strip_pcie_id(line)
                if id_data is not None:
                    pcie_useful_data_set.add(id_data)
            else:
                pcie_not_recognizable_data_list.append(line)
print_info("========================================")

# 对"无效信息"进行检查
for line in pcie_not_useful_data_set:
    if pcie_input_is_valid(line):
        pcie_useful_data_set.add(strip_pcie_id(line))

# 移除无关元素
try:
    pcie_useful_data_set.remove("")
    pcie_not_recognizable_data_list.remove("PCIe_PCI设备信息缺失项")
except KeyError:
    pass

# 查询该设备信息是否已经存在于pci.ids中, 如果是, 则记录
pci = PCI()
for device in pcie_useful_data_set:
    ret = pci.get_device(device[0:4], device[5:9])
    if ret:
        pcie_already_existed_data_set.add(device)

# 对两个集合取差集, 获得目前信息未知的设备编号
pcie_useful_data_set = pcie_useful_data_set.difference(pcie_already_existed_data_set)

# ==================== USB 部分 ====================

# 遍历列表,逐行判断是否为有效的 USB 数据
print_info("正在整理 USB 设备信息...")
for block in usb_raw_data_list:
    current_block = str(block.value).splitlines()
    for line in current_block:
        line = line.strip()
        if not line_is_useful(line):
            usb_not_useful_data_set.add(line)
        else:
            id_data = strip_usb_id(line)
            usb_useful_data_set.add(id_data)
print_info("========================================")

# 移除无关元素
try:
    usb_useful_data_set.remove("")
except KeyError:
    pass

# 查询该设备信息是否已经存在于usb.ids中, 如果是, 则记录
usb = USB()
for device in usb_useful_data_set:
    ret = usb.get_device(device[0:4], device[5:9])
    if ret:
        usb_already_existed_data_set.add(device)

# 对两个集合取差集, 获得目前信息未知的设备编号
usb_useful_data_set = usb_useful_data_set.difference(usb_already_existed_data_set)

# ==================== 汇总部分 ====================

# 按字母排序整理信息列表
pcie_useful_data_list = sorted(list(pcie_useful_data_set))
usb_useful_data_list = sorted(list(usb_useful_data_set))

# 保存到工作簿中:
hwdata_workbook = openpyxl.load_workbook(hwdata_excel_file)

try:
    del hwdata_workbook["检索的PCIe数据"]
    del hwdata_workbook["检索的USB数据"]
except KeyError:
    pass
try:
    del hwdata_workbook["整理后的PCIe数据"]
    del hwdata_workbook["整理后的USB数据"]
except KeyError:
    pass

sheet_count = len(hwdata_workbook.sheetnames)
hwdata_sorted_pcie_sheet = hwdata_workbook.create_sheet("整理后的PCIe数据", sheet_count)
hwdata_sorted_usb_sheet = hwdata_workbook.create_sheet("整理后的USB数据", sheet_count + 1)
for i in range(len(pcie_useful_data_list)):
    hwdata_sorted_pcie_sheet.cell(i + 1, 1, pcie_useful_data_list[i])
for i in range(len(usb_useful_data_list)):
    hwdata_sorted_usb_sheet.cell(i + 1, 1, usb_useful_data_list[i])

save_success = False

while not save_success:
    try:
        hwdata_workbook.save(hwdata_excel_file)
        save_success = True
    except PermissionError:
        print_warn("无法写入表格文件, 程序无法继续!\n" +
                   "请关闭表格文件 \"hwdata.xlsx\"后, 按 [E] 键重试." +
                   "或按 [Q] 键退出程序."
                   )
        print_warn("========================================")
        wait_for_user_input()

hwdata_workbook.close()

# 汇报结果
pcie_count_useful = len(pcie_useful_data_set)
pcie_count_total = pcie_count_useful + len(pcie_already_existed_data_set)

usb_count_useful = len(usb_useful_data_set)
usb_count_total = usb_count_useful + len(usb_already_existed_data_set)
print_info(f"整理完毕, 已将信息保存到表格中.")
print_info(f"收集表中有 {pcie_count_total} 条 PCI(e) 设备 ID, 其中有 {pcie_count_useful} 条是目前未知的.")
print_info(f"收集表中有 {usb_count_total} 条 USB 设备 ID, 其中有 {usb_count_useful} 条是目前未知的.")
print_info("========================================")
time.sleep(5)
clear_screen()
print_data("========================================")
print_data("注意: 还有一些内容无法识别, 请检查:")
time.sleep(1)
print_data("\nPCI(e) 部分:")
for line in pcie_not_recognizable_data_list:
    print_data(line)
print_data("\nUSB 部分:")
for line in usb_not_useful_data_set:
    print_data(line)
print_info("========================================")

time.sleep(1)
print_info(
    "接下来, 您需要检视和修改 \"hwdata.xlsx\", " +
    "确保所有的行都符合 [xxxx:xxxx] 的格式.\n" +
    "按下 [E] 键, 将为您打开表格. 您也可以按下 [Q] 键退出程序."
)

wait_for_user_input()
open_file(hwdata_excel_file)
time.sleep(2)

print_info("========================================")
print_info(
    "当您完成修改后，按下 [E] 键, 将继续执行 (请确保您已经关闭电子表格程序).\n" +
    "您也可以按下 [Q] 键退出程序."
)
print_info("========================================")
wait_for_user_input()

hwdata_workbook = openpyxl.load_workbook(hwdata_excel_file)
hwdata_pcie_sheet = hwdata_workbook["整理后的PCIe数据"]
hwdata_usb_sheet = hwdata_workbook["整理后的USB数据"]

pcie_data_list = list(hwdata_pcie_sheet.columns)[0]
usb_data_list = list(hwdata_usb_sheet.columns)[0]
hwdata_workbook.close()

# 创建字典(用于多线程)
pcie_name_on_linuxhardware_dict = dict()
pcie_name_on_driverscollection_dict = dict()
pcie_name_on_treexy_dict = dict()

usb_name_on_linuxhardware_dict = dict()
usb_name_on_driverscollection_dict = dict()
usb_name_on_treexy_dict = dict()

for block in pcie_data_list:
    pci_id = block.value
    pcie_name_on_linuxhardware_dict[pci_id] = "检索中..."
    pcie_name_on_driverscollection_dict[pci_id] = "检索中..."
    pcie_name_on_treexy_dict[pci_id] = "检索中..."

for block in usb_data_list:
    usb_id = block.value
    usb_name_on_linuxhardware_dict[usb_id] = "检索中..."
    usb_name_on_driverscollection_dict[usb_id] = "检索中..."
    usb_name_on_treexy_dict[usb_id] = "检索中..."


def search_pci_linuxhardware_and_save(_pci_id: str) -> None:
    result = search_pci_linuxhardware(_pci_id)
    pcie_name_on_linuxhardware_dict[_pci_id] = result


def search_pci_driverscollection_and_save(_pci_id: str) -> None:
    result = search_pci_driverscollection(_pci_id)
    pcie_name_on_driverscollection_dict[_pci_id] = result


def search_pci_treexy_and_save(_pci_id: str) -> None:
    result = search_pci_treexy(_pci_id)
    pcie_name_on_treexy_dict[_pci_id] = result


def search_usb_linuxhardware_and_save(_usb_id: str) -> None:
    result = search_usb_linuxhardware(_usb_id)
    usb_name_on_linuxhardware_dict[_usb_id] = result


def search_usb_driverscollection_and_save(_usb_id: str) -> None:
    result = search_usb_driverscollection(_usb_id)
    usb_name_on_driverscollection_dict[_usb_id] = result


def search_usb_treexy_and_save(_usb_id: str) -> None:
    result = search_usb_treexy(_usb_id)
    usb_name_on_treexy_dict[_usb_id] = result


clear_screen()
print_info("========================================")
# 开始检索
print_info(f"开始在线检索未知的设备名称, 这可能需要几分钟... ")
print_info("信息来源: LinuxHardware.org, DriversCollection.com, Treexy.com")
print_info("========================================")

# 最大线程数可根据实际情况修改
with ThreadPoolExecutor(max_workers=8) as executor:
    print_info("正在检索 PCI(e) 设备名称信息")
    for pci_id in pcie_name_on_linuxhardware_dict:
        executor.submit(search_pci_linuxhardware_and_save, pci_id)
        executor.submit(search_pci_driverscollection_and_save, pci_id)
        executor.submit(search_pci_treexy_and_save, pci_id)

    print_info("正在检索 USB 设备名称信息")
    for usb_id in usb_name_on_linuxhardware_dict:
        executor.submit(search_usb_linuxhardware_and_save, usb_id)
        executor.submit(search_usb_driverscollection_and_save, usb_id)
        executor.submit(search_usb_treexy_and_save, usb_id)

    executor.shutdown(wait=True)

print_info(f"已经完成全部检索!")

# 保存到工作簿中
hwdata_workbook = openpyxl.load_workbook(hwdata_excel_file)

try:
    del hwdata_workbook["检索的PCIe数据"]
    del hwdata_workbook["检索的USB数据"]
except KeyError:
    pass
try:
    del hwdata_workbook["整理后的PCIe数据"]
    del hwdata_workbook["整理后的USB数据"]
except KeyError:
    pass

sheet_count = len(hwdata_workbook.sheetnames)
hwdata_sorted_pcie_sheet = hwdata_workbook.create_sheet("检索的PCIe数据", sheet_count)
hwdata_sorted_usb_sheet = hwdata_workbook.create_sheet("检索的USB数据", sheet_count + 1)

hwdata_sorted_pcie_sheet.cell(1, 1, "PCI(e) 设备 ID")
hwdata_sorted_pcie_sheet.cell(1, 2, "LinuxHardware 查询结果")
hwdata_sorted_pcie_sheet.cell(1, 3, "DriversCollection 查询结果")
hwdata_sorted_pcie_sheet.cell(1, 4, "Treexy 查询结果")

hwdata_sorted_usb_sheet.cell(1, 1, "USB 设备 ID")
hwdata_sorted_usb_sheet.cell(1, 2, "LinuxHardware 查询结果")
hwdata_sorted_usb_sheet.cell(1, 3, "DriversCollection 查询结果")
hwdata_sorted_usb_sheet.cell(1, 4, "Treexy 查询结果")

pcie_id_list = list(pcie_name_on_linuxhardware_dict.keys())
usb_id_list = list(usb_name_on_linuxhardware_dict.keys())

pcie_name_on_linuxhardware_list = list(pcie_name_on_linuxhardware_dict.values())
pcie_name_on_driverscollection_list = list(pcie_name_on_driverscollection_dict.values())
pcie_name_on_treexy_list = list(pcie_name_on_treexy_dict.values())

usb_name_on_linuxhardware_list = list(usb_name_on_linuxhardware_dict.values())
usb_name_on_driverscollection_list = list(usb_name_on_driverscollection_dict.values())
usb_name_on_treexy_list = list(usb_name_on_treexy_dict.values())

for i in range(len(pcie_id_list)):
    hwdata_sorted_pcie_sheet.cell(i + 2, 1, pcie_id_list[i])
    hwdata_sorted_pcie_sheet.cell(i + 2, 2, pcie_name_on_linuxhardware_list[i])
    hwdata_sorted_pcie_sheet.cell(i + 2, 3, pcie_name_on_driverscollection_list[i])
    hwdata_sorted_pcie_sheet.cell(i + 2, 4, pcie_name_on_treexy_list[i])

for i in range(len(usb_id_list)):
    hwdata_sorted_usb_sheet.cell(i + 2, 1, usb_id_list[i])
    hwdata_sorted_usb_sheet.cell(i + 2, 2, usb_name_on_linuxhardware_list[i])
    hwdata_sorted_usb_sheet.cell(i + 2, 3, usb_name_on_driverscollection_list[i])
    hwdata_sorted_usb_sheet.cell(i + 2, 4, usb_name_on_treexy_list[i])

print_info("========================================")
save_success = False

while not save_success:
    try:
        hwdata_workbook.save(hwdata_excel_file)
        save_success = True
    except PermissionError:
        print_warn("无法写入表格文件, 程序无法继续!\n" +
                   "请关闭表格文件 \"hwdata.xlsx\"后, 按 [E] 键重试." +
                   "或按 [Q] 键退出程序."
                   )
        print_warn("========================================")
        wait_for_user_input()

hwdata_workbook.close()

print_info(f"已将在线检索的设备信息保存到表格中!")
print_info("========================================")
print_info("您可以按下 [E] 键打开电子表格检视结果, 也可以按下 [Q] 键中止程序.")
print_info("========================================")
wait_for_user_input()
open_file(hwdata_excel_file)
time.sleep(2)

print_info("已完成全部任务, 结束程序...")
